import openpyxl as xl
# importando classes para add gráficos

from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    # importando documento

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # usando loop para apresentar valores de uma coluna específica

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        # calculando valor por valor de cada célula da coluna
        corrected_price = cell.value * 0.9
        # add resultado em uma nova coluna
        corrected_price_sell = sheet.cell(row, 4)
        # add valor na célula
        corrected_price_sell.value = corrected_price

# add range de valores para pegar dados e inserir no gráfico

# Criando construtor

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # salvando documento

    wb.save(filename)

# Executando a função, usando o arquivo na pasta de projeto como parâmetro


process_workbook("transactions.xlsx")


